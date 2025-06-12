"""
Módulo para almacenar los resultados de cada intento del bot en tres formatos:
- CSV
- JSON
- XLSX

Cada resultado incluye:
- ID único (5 caracteres)
- Número de tarjeta usada
- Merchant
- Plan
- Mes de expiración
- Año de expiración
- CVV
- Tipo de tarjeta
- Banco
- Nombre de tarjeta
- Red de la tarjeta
- País
- Fecha y hora del intento
- Mensaje de estado
"""

import os
import csv
import json
from openpyxl import Workbook, load_workbook

# Directorio donde se almacenan los resultados
LOG_DIR = "logs"
CSV_FILE = os.path.join(LOG_DIR, "resultados.csv")
JSON_FILE = os.path.join(LOG_DIR, "resultados.json")
XLSX_FILE = os.path.join(LOG_DIR, "resultados.xlsx")

# Campos que se guardarán en todos los formatos
FIELDNAMES = [
    "id",
    "tarjeta",
    "merchant",
    "plan",
    "mes",
    "año",
    "cvv",
    "tipo",
    "banco",
    "nombre_tarjeta",
    "red",
    "pais",
    "fecha",
    "mensaje"
]

def guardar_resultado(data):
    """
    Guarda los datos del intento actual en los tres formatos.
    Args:
        data (dict): Diccionario con todos los campos necesarios
    """
    os.makedirs(LOG_DIR, exist_ok=True)
    escribir_csv(data)
    escribir_json(data)
    escribir_excel(data)

def escribir_csv(data):
    """
    Guarda los resultados en formato CSV.
    Crea el archivo si no existe.
    """
    existe = os.path.isfile(CSV_FILE)
    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        if not existe:
            writer.writeheader()
        writer.writerow(data)

def escribir_json(data):
    """
    Guarda los resultados en formato JSON como lista de objetos.
    """
    todos = []
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE, "r", encoding="utf-8") as f:
            try:
                todos = json.load(f)
            except json.JSONDecodeError:
                todos = []
    todos.append(data)
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(todos, f, indent=4, ensure_ascii=False)

def escribir_excel(data):
    """
    Guarda los resultados en una hoja de Excel.
    Crea el archivo y encabezados si no existe.
    """
    if not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(FIELDNAMES)  # Usar los mismos campos que CSV
    else:
        wb = load_workbook(XLSX_FILE)
        ws = wb["Resultados"]

    # Agregar datos en el mismo orden que los campos
    row_data = [data.get(field, "") for field in FIELDNAMES]
    ws.append(row_data)
    wb.save(XLSX_FILE)
